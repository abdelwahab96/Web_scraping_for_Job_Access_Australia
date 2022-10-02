#modules importing
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
import unittest

options = Options()
options.add_argument('--headless')
options.add_argument('--disable-gpu')
options.use_chromium = True 

service = Service(executable_path="c:\\users\\shaba7\\appdata\\local\\programs\\python\\python39\\msedgedriver.exe",verbose = True)
driver = webdriver.Edge(service=service, options=options) 

from bs4 import BeautifulSoup as bs
import requests
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl import Workbook
import xlsxwriter
import re
import datetime
import dropbox
now = datetime.datetime.now()



def operate(i,pathcode):
    lst_of_lnks = [] #list of links that cintains href for each single services
    nums = 18275
    
    
    postcodes_num = i
    path_code = pathcode

    print(f"------------------------- \nyou are scraping {pathcode} \npostcode number: {postcodes_num} \n-------------------------")
    #path_code = str(aus_codes.iloc[int(get_code_num)].values).replace("['", "").replace("']","")
    iters = 0
    
    
    path_url  = 'https://www.jobaccess.gov.au/find-a-provider?field_geofield_distance%5Borigin%5D='+path_code+'&field_geofield_distance%5Bdistance%5D=10&title=&field_service_value=1'
    code_page = str((path_code.split('+')[0])+"_"+(path_code.split('+')[1]))
    try:
        dff = pages_toggle(path_url, path_code, lst_of_lnks)

        print(dff)
        try:
            if dff == False:
                print('we r in first try')
                num_rows = 0
                gen_file(postcodes_num ,path_code, 'NO DATA', num_rows)
        
        except:    
            print('we r in second except')
            num_rows = len(dff)
            folder_path = get_scoial_myfunc(dff, path_code)
            gen_file(postcodes_num ,path_code, folder_path ,num_rows  )
    except Exception as e:
        print('we r in first except')
        print(e)
        num_rows = 0
        gen_file(postcodes_num ,path_code, 'NO DATA', num_rows)


            
    #dff = dff.replace(r'^\s*$', "NA", regex=True)
    #dff.to_excel("test2.xlsx",index=False)
    
    

    return print(f"------------------------- \nNow you finished {i}/{nums} \nscrapped {num_rows} rows \n-------------------------")







## scraping single page function 
def scrap():
    
    # for phone or free call 
    #use if to include both in phone list
    company_lst = []
    serv_lst =[]
    specialis_lst = []
    address_lst = []
    suburb_lst =[]
    state_lst =[]
    postcode_lst = []

    phone_lst = []
    web_lst = []
    email_lst=[]
    fbook_lst = []
    twit_lst = []
    yot_lst = []
    lnkd_lst = []
    insta_lst = []
    
    
    job_seekers_lst = []
    employers_lst = []
    networks_lst =[]
    about_us_lst = []

    ## inside each page
    page_box = driver.find_element(By.CSS_SELECTOR, '.block-main')
    #info_box
    info_box = page_box.find_element(By.CSS_SELECTOR, '.one_half')
    

    #get title of the company 
    com = page_box.find_element(By.CSS_SELECTOR, 'article h1')
    if com.text =="":
        company = "NA"
    else:    
        company = com.text

    #info box includes (service , Specialisation ,address ,Suburb ,state ,postcode ,Phone ,Email )
    info_box_child = info_box.find_elements(By.CSS_SELECTOR, '.provider-info-box')
    site_info = info_box_child[0]
    cont_info =info_box_child[1]

    all_p = site_info.find_elements(By.CSS_SELECTOR, 'div p')
    #string of the services
    serv_pr = all_p[0].text.split("\n")[1]
    if serv_pr =='':
        serv = 'NA'
    else: 
       serv = serv_pr
    #string of specialization
    speci = all_p[1].text.split('\n')[1]
    if speci == '':
        specialist ="NA"
    else:
        specialist = speci

    #get the address, suburb, post code
     #####################################################
    address_container = site_info.find_element(By.CSS_SELECTOR, '.addressfield-container-inline')
    if len(address_container.text.split(" ")) >=4:
        suburb_pr = (address_container.text.split(" ")[0]) +" "+ (address_container.text.split(" ")[1])
        if suburb_pr == '':
            suburb = "NA"
        else:
            suburb = suburb_pr

        state_pr = address_container.text.split(" ")
        state_pri = state_pr[len(state_pr)-2]
        if state_pri == "":
            state = "NA"
        else:
            state = state_pri
    else:
        suburb_pr = address_container.text.split(" ")[0]
        if suburb_pr == '':
            suburb = "NA"
        else:
            suburb = suburb_pr

        state_pr = address_container.text.split(" ")
        state_pri = state_pr[len(state_pr)-2]
        if state_pri == "":
            state = "NA"
        else:
            state = state_pri

    ################################
    post_code_pr = address_container.text.split(" ")
    post_code_pri = post_code_pr[len(post_code_pr)-1]
    if post_code_pri =="":
        postcode = "NA"
    else:
        postcode  = post_code_pri

    add_street = site_info.find_element(By.CSS_SELECTOR,".street-block")
    address_pr = add_street.text+" "+address_container.text
    if address_pr == '':
        address = "NA"
    else:
        address = address_pr
    #print(add_street.text+" "+address_container.text)

    ##contact info
    cont_info =info_box_child[1]
    lst_contacts = cont_info.find_elements(By.CSS_SELECTOR, 'div p')


    

    # 'PR' for primary
    phone_pr = ''
    freecall_pr = ''
    web_pr = ''
    email_pr = ''
    fbook_pr = ''
    twit_pr = ''
    tube_pr = ''
    linkd_pr = ''
    insta_pr = ''
    tk = ''

    for i in lst_contacts:
        
        if i.text.split(":")[0] == 'Phone'  :
            phone_pr = i.text.split(":")[1]
            continue
        
        elif i.text.split(":")[0] == 'Freecall':
            freecall_pr = i.text.split(":")[1]
        
        
        elif i.text.split(":")[0] == 'Website' :
            web_pr= i.text.split(": ")[1]
            continue
        elif i.text.split(":")[0] == 'Email':
            email_pr = i.text.split(":")[1]
            continue
        
        elif i.text.split(":")[0] == 'Facebook':
            f_lnk = i.find_element(By.CSS_SELECTOR, "a").get_attribute('href')
            fbook_pr = f_lnk
            continue
        
        elif i.text.split(":")[0] == 'Twitter':
            twt_lnk = i.find_element(By.CSS_SELECTOR, "a").get_attribute('href')
            twit_pr = twt_lnk
            continue
        
        elif i.text.split(":")[0] == "YouTube":
            yt_lnk = i.find_element(By.CSS_SELECTOR, "a").get_attribute('href')
            tube_pr = yt_lnk
            continue
        elif i.text.split(":")[0] == "LinkedIn" or i.text.split(":")[0] == 'Linkedin':
            l_lnk = i.find_element(By.CSS_SELECTOR, "a").get_attribute('href')
            linkd_pr = l_lnk
            continue
        
        elif i.text.split(":")[0] == "Instagram":
            inst_lnk = i.find_element(By.CSS_SELECTOR, "a").get_attribute('href')
            insta_pr = inst_lnk
            continue
        
            

    if phone_pr == "":
        phone = "NA"
    else:
        phone = phone_pr

    if freecall_pr != '':
        phone = phone_pr+", "+freecall_pr 
    else:
        pass

    if web_pr == '':
        web= "NA"
    else:
        web = web_pr
        
    if email_pr == '':
        email= "NA"
    else:
        email = email_pr
        
    if fbook_pr == '':
        fbook = "NA"
    else:
        fbook = fbook_pr

    if twit_pr == '':
        twit = "NA"
    else:
        twit = twit_pr

    if tube_pr =="":
        yot = "NA"
    else:
        yot = tube_pr

    if linkd_pr == '':
        lnkd = 'NA'
    else:
        lnkd = linkd_pr

    if insta_pr == "":
        insta = "NA"
    else:
        insta = insta_pr


    other_info_box = page_box.find_elements(By.CSS_SELECTOR, "article > div.provider-info-box") 
    #desired_lst = ['Our Job Seekers','Our Employers','Our Networks','About Us']

    parag = ''
    for box in other_info_box:
        try:
            box_text = box.find_element(By.CSS_SELECTOR, "h2").text
            if box_text == 'Connections for Quality':
            
                every_elements = box.find_elements(By.CSS_SELECTOR, "*") 
                h2_element = box.find_element(By.TAG_NAME, 'h3').text ## every heading (oursss)
                for ind, c in enumerate(every_elements):
                    
                    if ind ==1:
                        parag = str(ind)+" "+c.text
                        break
                
            else:
                continue
                
        except:
            print('something wrong is done in scrap function ')
            continue
    job_seekers = parag.split('Our Job Seekers')[1].split('Our Employers')[0].strip('\n').replace('\n','')
    #job_seekers_lst.append(str(job_seekers))
    employers = parag.split('Our Employers')[1].split('Our Networks')[0].strip('\n').replace('\n','')
    #employers_lst.append(employers)
    networks = parag.split('Our Networks')[1].split('Our Performance')[0].strip('\n').replace('\n','')
    #networks_lst.append(networks)

    about_us = parag.split('About Us')[1].strip('\n').replace('\n','')
    #about_us_lst.append(about_us)

    
    return (company, serv , specialist ,address , suburb, state, postcode,phone,web,email,fbook, twit, yot, lnkd, insta , tk ,job_seekers,employers,networks,about_us)
############## end of scraping sigle page function  #############################


##### start toggling between pages function #####
    
def pages_toggle(path,  code_page, lst_of_lnks):
    inner_lst_of_lnks = lst_of_lnks
    path = path
    code_path = path.split('=')[1].split('&')[0].replace("%",'+')
    code_page = code_page
    #print(code_page)
    #inner_iters = iters +1
    driver.get(path)
    time.sleep(2)
    lsting =[]
    #new_lst = []
    try:
        df_testing = pd.DataFrame(columns=["Company","Services" ,"Specialization" ,'Address' , "Suburb","State", "Postcode" ,"Contacts" ,"Website" ,"Email","Facebook", "Twitter","YouTube"  , "Linkedin","Instagram", "Tiktok", "Job Seekers",'Employers',"Networks", "About us"])
    except:
        pass
    time.sleep(3)
    #create a folder using code page
    #try:
        #os.mkdir("C:\\Users\\SHABA7\\DA-py-libs\\SEEKA\\Job_access\\single_pages\\"+code_page)
        #print('folder_created')
    #except:
        #print('already created')

    
    try:    
        holder = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, '.views-unformatted')))
        
        services_children = holder.find_elements(By.CSS_SELECTOR, 'div.provider-result h3 a')
        children_lnks = [ch.get_attribute('href') for ch in services_children]
        
        for c in children_lnks:
            inner_lst_of_lnks.append(str(c))
        
        driver.get(path) #back to main page

        
        time.sleep(2)
        try:
            next_but = driver.find_element(By.CSS_SELECTOR, 'li.pager-next  a').get_attribute('href')
            pages_toggle(next_but,  code_page, inner_lst_of_lnks )
        
        except:
            pass
        
        for each_serv in inner_lst_of_lnks:
            driver.get(each_serv)
            time.sleep(2)
            lsting = list(scrap())
            df_testing.loc[len(df_testing)+1] = lsting
        
        print('we r in holder ', len(df_testing))
        
    except:
        df_testing = False
        print(df_testing)
    
    inner_lst_of_lnks = []  
    return df_testing
##### end toggling between pages function #####


## scraping social media from the website using regex
def get_scoial_myfunc(df, code_page):
    code_page = code_page
    #lists for each link
    web_lst= []
    fb_lst =[]
    inst_lst= []
    tk_lst = []
    lnk_lst =[]
    twt_lst =[]
    email_miss_lst = []
    
    
    
    
    #get the website
    facebook_df = list(df['Facebook'])
    website = list(df['Website'])
    #print(len(website))
    for x,i in enumerate(website):
        #print(x)
        if str(i) == 'nan':
            #print("empty web")
            web_lst.append("NA")
            if pd.isna(df['Facebook'].iloc[x]) :
                
                fb_lst.append("NA")
                #print("ooh it's reaaaly here: ",str(facebook_df[x]))
                
            else:
                #print('here, the empty facebook is:', str(facebook_df[x]))
                fb_lst.append(df['Facebook'].iloc[x])

            if pd.isna(df['Instagram'].iloc[x]) :
                inst_lst.append("NA")
            else:
                inst_lst.append(df['Instagram'].iloc[x])

            
            tk_lst.append('NA')

            if pd.isna(df['Linkedin'].iloc[x]):
                lnk_lst.append('NA')
            else:
                lnk_lst.append(df['Linkedin'].iloc[x])
            
            if pd.isna(df['Twitter'].iloc[x]):
                twt_lst.append('NA')
            else:
                twt_lst.append(df['Twitter'].iloc[x])
            
        else:
            f = re.compile("https?://(www\.)?facebook\.com/(?!share\.php).(\S+\.?)+")  # regex for facebook
            matching = f.match(str(i))
            
            if bool(matching) is True:
                fb_lst.append(i)
                web_lst.append('NA')
                if pd.isna(df['Instagram'].iloc[x]) :
                    inst_lst.append("NA")
                else:
                    inst_lst.append(df['Instagram'].iloc[x])
                tk_lst.append('NA')

                if pd.isna(df['Linkedin'].iloc[x]):
                    lnk_lst.append('NA')
                else:
                    lnk_lst.append(df['Linkedin'].iloc[x])

                if pd.isna(df['Twitter'].iloc[x]):
                    twt_lst.append('NA')
                else:
                    twt_lst.append(df['Twitter'].iloc[x])


            else:
                website = i
                web_lst.append(website)
                
                for ss in range(3):
                    try:
                        html_data = None
                        check = None
                        if ss == 0:
                            try:
                                driver.get(str(website))
                                #print(str(website))
                                driver.set_page_load_timeout(10)
                                html_data = driver.page_source
                            except:
                                driver.get("https://"+str(website))
                                #print("https://"+str(website))
                                driver.set_page_load_timeout(10)
                                html_data = driver.page_source
                                
                    
                        bsoup = bs(html_data, 'html.parser')
                        #if bsoup is None:
                         #   print("the problem is here")
                        fb = bsoup.find('a', {'href': re.compile("https?://(www\.)?facebook\.com/(?!share\.php).(\S+\.?)+")})
                        if fb is None:
                            facebook = 'NA'
                            #print('cant find fb')
                        else:
                            facebook = fb['href']

                        #get email 
                        em = bsoup.find('a', {'href':re.compile(r"(?:[a-z0-9!#$%&'*+/=?^_`{|}~-]+(?:\.[a-z0-9!#$%&'*+/=?^_`{|}~-]+)*|\"(?:[\x01-\x08\x0b\x0c\x0e-\x1f\x21\x23-\x5b\x5d-\x7f]|\\[\x01-\x09\x0b\x0c\x0e-\x7f])*\")@(?:(?:[a-z0-9](?:[a-z0-9-]*[a-z0-9])?\.)+[a-z0-9](?:[a-z0-9-]*[a-z0-9])?|\[(?:(?:(2(5[0-5]|[0-4][0-9])|1[0-9][0-9]|[1-9]?[0-9]))\.){3}(?:(2(5[0-5]|[0-4][0-9])|1[0-9][0-9]|[1-9]?[0-9])|[a-z0-9-]*[a-z0-9]:(?:[\x01-\x08\x0b\x0c\x0e-\x1f\x21-\x5a\x53-\x7f]|\\[\x01-\x09\x0b\x0c\x0e-\x7f])+)\])")})
                        if em is None:   
                            ema = 'NA'
                        else:
                            ema = em['href']

                        ig = bsoup.find('a',
                                    {'href': re.compile(
                                        "https?://(www\.)?instagram\.com/(?!share\.php).(\S+\.?)+")})
                        if ig is None:
                            instagram = 'NA'
                        else:
                            instagram = ig['href']

                        tt = bsoup.find('a',
                                    {'href': re.compile(
                                    "https?://(www\.)?tiktok\.com/(?!share\.php).(\S+\.?)+")})
                        if tt is None:
                            tiktok = 'NA'
                        else:
                            tiktok = tt['href']

                        li = bsoup.find('a',
                                {'href': re.compile(
                                        "https?://(www\.)?linkedin\.com/(?!share\.php).(\S+\.?)+")})
                        if li is None:
                            linkedin = 'NA'
                        else:
                            linkedin = li['href']

                        tw = bsoup.find('a',
                                    {'href': re.compile(
                                        "https?://(www\.)?twitter\.com/(?!share\.php).(\S+\.?)+")})
                        if tw is None:
                            twitter = 'NA'
                        else:
                            twitter = tw['href']
                        break
                    except:
                        #print(f"Can't reach the website.\n{website}")
                        time.sleep(1)
                        facebook = 'NA'
                        instagram = 'NA'
                        tiktok = 'NA'
                        linkedin = 'NA'
                        twitter = 'NA'
                        continue
                
                if pd.isna(df['Facebook'].iloc[x]) :
                    fb_lst.append(facebook)   
                else:    
                    fb_lst.append(df['Facebook'].iloc[x])

                if pd.isna(df['Instagram'].iloc[x]) :
                    inst_lst.append(instagram)
                else:
                    inst_lst.append(df['Instagram'].iloc[x])


                if pd.isna(df['Linkedin'].iloc[x]):
                    lnk_lst.append(linkedin)
                else:
                    lnk_lst.append(df['Linkedin'].iloc[x])
                
                if pd.isna(df['Twitter'].iloc[x]):
                    twt_lst.append(twitter)
                else:
                    twt_lst.append(df['Twitter'].iloc[x])

                if pd.isna(df['Email'].iloc[x]):
                    email_miss_lst.append(ema)
                else:
                    email_miss_lst.append(df['Email'].iloc[x])
                    
                tk_lst.append(tiktok)
                
                
                                
            
    
 
    df['Website']=  web_lst
    df['Facebook']=  fb_lst
    df['Instagram']=  inst_lst
    df['Tiktok']=  tk_lst
    df['Linkedin']=  lnk_lst
    df['Twitter']=  twt_lst
    df['Email'] = get_email_facebook(df)
   
    
        
        
        
    #pd.to_excel()
    try:
        #df.drop_duplicates(keep='first', inplace=True)
        df = df.replace(r'^\s*$', "NA", regex=True)
        code_page = code_page.replace('+','_')
        df.to_excel(os.getcwd()+"\\job_access_official\\Job_Access_DES_"+code_page+".xlsx" , index=False,na_rep='NA')
    except:
        #df.drop_duplicates(keep='first', inplace=True)
        df = df.replace(r'^\s*$', "NA", regex=True)
        df.to_excel("job_acc_social_error_1.xlsx", index=False,na_rep='NA')
        
    folder_path = os.getcwd()+"\\job_access_official\\Job_Access_DES_"+code_page+".xlsx"
    return  folder_path



## end social media scraping ###

##combining function and exporting as one excel file
def combine_df():
    
    excel_files_path = os.path.join( os.getcwd() , "job_access_official")
    all_files = os.listdir(excel_files_path)
    files_path = [os.path.join(excel_files_path, f) for f in all_files ]

    combined_df = pd.concat([pd.read_excel(f_path) for f_path in files_path] , ignore_index=True)
    combined_df.drop_duplicates(keep='first', inplace=True)
    combined_df = combined_df.replace(r'^\s*$', "NA", regex=True)
    #combined_df.columns=["Company","Services" ,"Specialization" ,'Address' , "Suburb","State", "Postcode" ,"Contacts" ,"Website" ,"Email","Facebook", "Twitter","YouTube" , "Linkedin","Instagram", "Tiktok","Job Seekers",'Employers',"Networks", "About us"]
    combined_df.to_excel("job_access_final_file.xlsx",index=False,na_rep='NA')
    return print("congrats! you finished scraping the whole website")

### end combining function ####

def get_path(path, usein, driver, ttl=30):
    WebDriverWait(driver, ttl).until(EC.presence_of_all_elements_located((usein, path)))
    return driver.find_element(by=usein, value=path)

######### start email from facebook function ###########
def get_email_facebook(df):

    facebook = list(df['Facebook'])
    email_lst = list(df['Email'])
    emails =[]

    for j , y in enumerate(email_lst):
        if pd.isna(df['Email'].iloc[j]) or df['Email'].iloc[j] == 'NA' or df['Email'].iloc[j] =='nan':
            fbok = df['Facebook'].iloc[j]
            if pd.isna(str(df['Facebook'].iloc[j])) or str(fbok) == 'nan' or str(fbok) == '' or str(fbok) =='NA':
                
                emails.append('NA')
                continue
            else:
                
                try:
                    driver.get(str(fbok))
                    
                    time.sleep(3)
                    email_rex = r'''(?:[a-z0-9!#$%&'*+/=?^_`{|}~-]+(?:\.[a-z0-9!#$%&'*+/=?^_`{|}~-]+)*|\"(?:[\x01-\x08\x0b\x0c\x0e-\x1f\x21\x23-\x5b\x5d-\x7f]|\\[\x01-\x09\x0b\x0c\x0e-\x7f])*\")@(?:(?:[a-z0-9](?:[a-z0-9-]*[a-z0-9])?\.)+[a-z0-9](?:[a-z0-9-]*[a-z0-9])?|\[(?:(?:(2(5[0-5]|[0-4][0-9])|1[0-9][0-9]|[1-9]?[0-9]))\.){3}(?:(2(5[0-5]|[0-4][0-9])|1[0-9][0-9]|[1-9]?[0-9])|[a-z0-9-]*[a-z0-9]:(?:[\x01-\x08\x0b\x0c\x0e-\x1f\x21-\x5a\x53-\x7f]|\\[\x01-\x09\x0b\x0c\x0e-\x7f])+)\])'''
                    Email = get_path("//body", By.XPATH, driver, 20).get_attribute('innerHTML')
                    final_email = re.search(email_rex, Email).group(0)
                    
                    if "login" in final_email or "profile" in final_email:
                        emails.append("NA")
                    else:
                        emails.append(final_email)
                except:
                    
                    final_email = "NA"
                    emails.append(final_email)
        else:
            emails.append(df['Email'].iloc[j])
            
            
    
    return emails

################### end email from facebook function #######




########### function of generating excel file
def gen_file(i, pathcode, pathurl,rows):
    index = i
    code_postcode = pathcode
    pathcode = code_postcode.split("+")
    postcode = pathcode.pop()
    state = pathcode.pop()
    suburb = " ".join(pathcode)
    url = pathurl
    status = 'Scrapped Successfully'
    scraped_rows= rows
    if scraped_rows == 0:
        url = 'NO DATA'
        status = 'NO DATA'
    
    scraped_time = now.strftime("%Y-%m-%d %H:%M:%S")
    try:
        ##read and write
        wb = load_workbook('job_access_diagnostic.xlsx')
        ws = wb.active
        
        ws.append([index,suburb,state,postcode,code_postcode.replace("+"," "),url,scraped_time,scraped_rows,status ])
        wb.save(filename='job_access_diagnostic.xlsx')
    except:
        headers  = ['index','suburb','state','postcode','url code','folder path','scrapped date','number of services', 'status']
        workbook_name = 'job_access_diagnostic.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        ws.append([index,suburb,state,postcode,code_postcode.replace("+"," "), url,scraped_time,scraped_rows,status ])
        wb.save(workbook_name)
    return





## put your starting function below
def start_scrap():
    
    aus_codes = pd.read_excel('aus_postcodes.xlsx',header=0)
    try:
        generator_file = pd.read_excel("job_access_diagnostic.xlsx")
        latest_row = generator_file['url code'].loc[len(generator_file['url code'])-1]
        row_length = len(generator_file)
    except:
        row_length = 0

    if row_length == 0:
        
        for i,x in enumerate(aus_codes['searching code']):

            try:
                operate(i+1,x)
                start_scrap()
            except Exception as e: 
                print(f"------------------------ \nyou got an exception: '{e}' \nin the postcode {x} \nand its index is {i+1} \n------------------------ ")
                continue
    elif row_length > 0:
        
        
        for i,x in enumerate(aus_codes['searching code']):
            time.sleep(3)
            generator_file = pd.read_excel("job_access_diagnostic.xlsx")
            latest_row = generator_file['url code'].loc[len(generator_file['url code'])-1]

            if x == latest_row.replace(" ","+"):
                operate(i+2,aus_codes['searching code'].loc[i+1] )  

            else:
                continue

    return 

start_scrap()
try:
    driver.dispose()
    #print("disposed")
except:
    driver.quit()
    print('quitted')

