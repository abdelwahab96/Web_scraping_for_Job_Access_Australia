
#modules importing
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
import unittest

options = Options()
options.add_argument('--headless')
options.add_argument('--disable-gpu')
options.use_chromium = True 

service = Service(executable_path="c:\\users\\shaba7\\appdata\\local\\programs\\python\\python39\\msedgedriver.exe",verbose = True)
driver = webdriver.Edge(service=service, options=options) 
from bs4 import BeautifulSoup as bs
import requests
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl import Workbook
import xlsxwriter
import re
import datetime

import dropbox
from tqdm import tqdm
now = datetime.datetime.now()

def connect_dropbox():
     #you shoul create dropbox application and change these 2 variables with yours 
    app_key = "s24bbl4mgrl3d29"
    app_secret = "dcyl3vmlu40jk5j"
    auth_flow = dropbox.DropboxOAuth2FlowNoRedirect(app_key, app_secret)

    auth_url = auth_flow.start()
    
    try:
        auth_code = get_auth_link(auth_url)
    except:
        print()
        print("\33[33m1. Go to: " + auth_url)
        print("2. Click \"Allow\" (you might have to log in first).")
        print("3. Copy the authorization code.\33[0m")
        auth_code = input("\33[4;33mEnter the authorization code here: \33[0m").strip()
    
    try:
        oauth_result = auth_flow.finish(auth_code)
    except Exception as e:
        print('Error: %s' % (e,))
        exit(1)
    with dropbox.Dropbox(oauth2_access_token=oauth_result.access_token) as dbx:
        print("Successfully set up client!")
        print()
        dbxs = dbx
    return dbxs


####### function of making shareable link
def get_drop_link(file_path,dbx,status):
    dbx =dbx
    file_path = file_path.replace(" ","_")
    status = status
    if status == 'NO DATA':
        shareble_link = 'NO DATA'
    else:
        #this path need to be changed
        shared_link_metadata = dbx.sharing_create_shared_link("/Abdelwahab Ahmed/03 My work/Job_Access/job_access_official/Job_Access_DES_"+file_path+".xlsx")
        shared_link = shared_link_metadata.url
        shareble_link = shared_link
    
    return shareble_link 

#####function of getting authorization link from dropbox
def get_auth_link(auth_url):
    driver = webdriver.Edge(service=service, options=options) 
    driver.get(auth_url)
    try:
        holder = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".credentials-form__fields")))
        try:
            children = holder.find_elements(By.CSS_SELECTOR, "input[type]")
            for i,ch in enumerate(children):
                if ch.get_attribute("type") == 'email':
                    ch.send_keys("abdelwahab@freshfutures.com")
                elif ch.get_attribute("type") == 'password':
                    ch.send_keys("yuzee22")
        except:
            ch_email= driver.find_element(By.NAME, 'login_email').send_keys("abdelwahab@freshfutures.com")
            ch_pass = driver.find_element(By.NAME, 'login_password').send_keys("yuzee22")
    

        try:
            button = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'button .signin-text')))
            button.click()
        except:
            driver.get("https://www.dropbox.com/oauth2/authorize?response_type=code&client_id=jo4g9e15ynnmocd")
            button = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'form button .signin-text')))
            button.click()
    except:
        pass

    try:
        cont_but = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.ID, 'warning-button-continue')))
        cont_but.click() 
    except:
        cont_but = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, '.app-warning-frame button')))
        for i,cont in enumerate(cont_but):
            if cont.get_attribute('id') == 'warning-button-continue':
                cont.click()

    try:
        allow_but = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.NAME, 'allow_access')))
        allow_but.click()
    except:
        butons = driver.find_elements(By.CSS_SELECTOR, '#buttons button')
        for i,b in enumerate(butons):
            if b.text == 'Allow':
                b.click()

    link_holder = driver.find_element(By.CSS_SELECTOR, '.auth-connect-scoped-frame input.auth-box')
    return link_holder.get_attribute("data-token")

def generate_file(shareable_lnk,diagnostic):
    sh_lnk = shareable_lnk
    df = diagnostic
    index = df['index']
    
    suburb = df['suburb']
    
    state = df['state']
    postcode = df['postcode']
    code_postcode = df['url code']
    scraped_time = df['scrapped date']
    scraped_rows = df['number of services']
    status = df['status']

    try:
        ##read and write
        wb = load_workbook('job_access_generator.xlsx')
        ws = wb.active
        
        ws.append([index,suburb,state,postcode,code_postcode,scraped_time,scraped_rows,status,sh_lnk ])
        wb.save(filename='job_access_generator.xlsx')
    except:
        headers  = ['index','suburb','state','postcode','url code','scrapped date','number of services', 'status','dropbox link']
        workbook_name = 'job_access_generator.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        ws.append([index,suburb,state,postcode,code_postcode,scraped_time,scraped_rows,status,sh_lnk  ])
        wb.save(workbook_name)
    return 

dbx = connect_dropbox()
def operate(dbx):
    dbx = dbx

    diagnostic= pd.read_excel("job_access_diagnostic.xlsx")
    status = diagnostic['status']
    try:
        generator_file = pd.read_excel("job_access_generator.xlsx")
        latest_row = generator_file['url code'].loc[len(generator_file['url code'])-1]
        row_length = len(generator_file)
    except:
        row_length = 0
    print(f"-------------------------------------------- \n>>>> you have {row_length}/{len(status)} shareable link <<<< \n--------------------------------------------")
    if row_length == 0:
            try:
                if status[0] == 'Scrapped Successfully':
                    
                    shareable_lnk = get_drop_link(diagnostic['url code'].loc[0],dbx,diagnostic['status'].loc[0])
                    generate_file(shareable_lnk,diagnostic.loc[0])
                    operate(dbx)
                    
                    
                elif status[0] =='NO DATA':
                    shareable_lnk = 'NO DATA'
                    generate_file(shareable_lnk,diagnostic.loc[0])
                    operate(dbx)

            except Exception as e: 
                print(f"------------------------ \nyou got an exception: '{e}' \nits index is 0 \n------------------------ ")
                
        

    elif row_length > 0:
        for i,x in enumerate(diagnostic['url code']):
            generator_file = pd.read_excel("job_access_generator.xlsx")
            latest_row = generator_file['url code'].loc[len(generator_file['url code'])-1]
            if i == len(diagnostic['url code'])-1 :
                break
            if x == latest_row :
                shareable_lnk = get_drop_link(diagnostic['url code'].loc[i+1],dbx,diagnostic['status'].loc[i+1])
                generate_file(shareable_lnk,diagnostic.loc[i+1])
            
            else:
                continue

    for i in tqdm ( range(len(status)), desc=f"generating {len(status) - row_length} files",ascii=False,ncols=75):
        time.sleep(0.01)
    return 

    

operate(dbx)
